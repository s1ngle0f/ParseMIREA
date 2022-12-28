import os
import pprint

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
import process_the_file
from process_the_file import process, get_by, process_dirty, process_exams
import copy_sheet
import DateTime


weekdays = {
    'ПОНЕДЕЛЬНИК': 0,
    'ВТОРНИК': 1,
    'СРЕДА': 2,
    'ЧЕТВЕРГ': 3,
    'ПЯТНИЦА': 4,
    'СУББОТА': 5,
    'ВОСКРЕСЕНЬЕ': 6
}

parity = {
    'I': 0,
    'II': 1
}

ban_symbols = [',', '\n\n', '\n', ', ']

parity_num_time = {
    1: '9-00',
    2: '10-40',
    3: '12-40',
    4: '14-20',
    5: '16-20',
    6: '18-00',
    7: '19-40'
}

def get_parity(time):
    try:
        if type(time) != float:
            # print(time)
            time = DateTime.DateTime(time)
            for k, v in parity_num_time.items():
                if time <= DateTime.DateTime(v):
                    if k > 1:
                        if DateTime.DateTime(v) - time < time - DateTime.DateTime(parity_num_time.get(k-1)):
                            return (k, v)
                        else:
                            return (k-1, parity_num_time.get(k-1))
                    else:
                        if time - DateTime.DateTime(v) < DateTime.DateTime(parity_num_time.get(k+1)) - time:
                            return (k, v)
                        else:
                            return (k+1, parity_num_time.get(k+1))
        else:
            return (8, 'Неизвестно')
    except:
        return (8, time)

def write(sort_by, name, result_filename):
    # preparated_data = get_by(sort_by, process(name))
    preparated_data = get_by(sort_by, process_dirty(name))
    try:
        wb = openpyxl.load_workbook(f'output/{result_filename}.xlsx')
    except:
        wb = openpyxl.Workbook()

        # Удаление листа, создаваемого по умолчанию, при создании документа
        for sheet_name in wb.sheetnames:
            sheet = wb.get_sheet_by_name(sheet_name)
            wb.remove_sheet(sheet)

    template = openpyxl.load_workbook('templates/pattern_prepod.xlsx')['Шаблон']

    for key, val in preparated_data.items():
        k = key.replace('/', ' ')
        try:
            ws = wb.get_sheet_by_name(k)
        except:
            ws = wb.create_sheet(k)
            copy_sheet.copy_sheet(template, ws)

        for info in val:
            # print(k, weekday, num_lesson, num_chetnost, info)
            line_num = 4+weekdays.get(info.get('День недели'))*16+(int(info.get('Номер пары'))-1)*2+parity.get(info.get('Четность недели'))
            ws[f'E{line_num}'] = info.get('Предмет')
            ws[f'F{line_num}'] = info.get('Вид занятий')
            ws[f'G{line_num}'] = info.get('ФИО')
            ws[f'H{line_num}'] = info.get('Аудитория')
            ws[f'I{line_num}'] = info.get('Группа')

            if any(ban_symbol in info.get('ФИО') for ban_symbol in ban_symbols):#Только для определения коллизий ФИО
                ws[f'J{line_num}'].fill = PatternFill(start_color="ff0000", end_color="ff0000", fill_type="solid")

    wb.save(f'output/{result_filename}.xlsx')

def write_exams(sort_by, name, result_filename):
    # preparated_data = get_by(sort_by, process(name))
    preparated_data = get_by(sort_by, process_exams(name))
    try:
        wb = openpyxl.load_workbook(f'output/{result_filename}.xlsx')
    except:
        wb = openpyxl.Workbook()

        # Удаление листа, создаваемого по умолчанию, при создании документа
        for sheet_name in wb.sheetnames:
            sheet = wb.get_sheet_by_name(sheet_name)
            wb.remove_sheet(sheet)

    template = openpyxl.load_workbook('templates/pattern_prepod_exams.xlsx')['Шаблон']

    for key, val in preparated_data.items():
        k = key.replace('/', ' ')
        try:
            ws = wb.get_sheet_by_name(k)
        except:
            ws = wb.create_sheet(k)
            copy_sheet.copy_sheet(template, ws)

        for info in val:
            # print(k, weekday, num_lesson, num_chetnost, info)
            # print(k)
            # print(info)
            # print(get_parity(info.get('Время')))
            # print(ws['Z20'].value)
            # print(type(ws['K7']))
            if info.get('Предмет').find('\n') == -1:
                line_num = 4+(int(info.get('День')) - 9)*16+(get_parity(info.get('Время'))[0]-1)*2
                ws[f'C{line_num}'] = info.get('Время')
                ws[f'D{line_num}'] = info.get('Предмет')
                ws[f'E{line_num}'] = info.get('Вид занятий')
                ws[f'F{line_num}'] = info.get('ФИО')
                ws[f'G{line_num}'] = info.get('Аудитория')
                if ws[f'H{line_num}'].value == None:
                    ws[f'H{line_num}'] = info.get('Группа')
                else:
                    ws[f'H{line_num}'] = ws[f'H{line_num}'].value + '\n' + info.get('Группа')
            else:
                print(info)
                predmets = info.get('Предмет').split('\n')
                for i, predmet in enumerate(predmets):
                    if type(info.get('Время')) != float:
                        _time = info.get('Время').split('\n')
                        line_num = 4 + (int(info.get('День')) - 9) * 16 + (get_parity(_time[i])[0] - 1) * 2
                    else:
                        line_num = 4 + (int(info.get('День')) - 9) * 16 + (7) * 2
                    if type(info.get('Аудитория')) != float:
                        _auditories = info.get('Аудитория').split('\n')
                        ws[f'G{line_num}'] = _auditories[i]
                    else:
                        _auditories = 'Неизвестно'
                        ws[f'G{line_num}'] = _auditories
                    ws[f'C{line_num}'] = info.get('Время')
                    ws[f'D{line_num}'] = predmet
                    ws[f'E{line_num}'] = info.get('Вид занятий')
                    if type(info.get('ФИО')) != float:
                        ws[f'F{line_num}'] = info.get('ФИО').split('\n')[i]
                    else:
                        ws[f'F{line_num}'] = 'Неизвестно'
                    ws[f'G{line_num}'] = _auditories[i]
                    if ws[f'H{line_num}'].value == None:
                        ws[f'H{line_num}'] = info.get('Группа')
                    else:
                        ws[f'H{line_num}'] = ws[f'H{line_num}'].value + '\n' + info.get('Группа')

    wb.save(f'output/{result_filename}.xlsx')

# write('IIT_3-kurs_22_23_osen_07.10.2022.xlsx')


def write_all_IiPPO_auditory(sort_by, result_filename):
    write_IiPPO_auditory(sort_by, os.listdir('input'), result_filename)

    def get_sheets_title(sheet):
        return sheet.title

    none_sort_wb = openpyxl.load_workbook(f'output/{result_filename}.xlsx')
    none_sort_wb._sheets.sort(key=get_sheets_title)
    none_sort_wb.save(f'output/{result_filename}.xlsx')


def write_all_exams(sort_by, result_filename):
    for files in os.listdir('input'):
        if files.find('ITU_mag') != -1:
            process_the_file.COUNT_GROUPS_IN_ONE_PART = 3
            write_exams(sort_by, files, result_filename)
        else:
            process_the_file.COUNT_GROUPS_IN_ONE_PART = 2
            write_exams(sort_by, files, result_filename)

    def get_sheets_title(sheet):
        return sheet.title

    none_sort_wb = openpyxl.load_workbook(f'output/{result_filename}.xlsx')
    none_sort_wb._sheets.sort(key=get_sheets_title)
    none_sort_wb.save(f'output/{result_filename}.xlsx')

def write_all(sort_by, result_filename):
    for files in os.listdir('input'):
        if files.find('ITU_mag') != -1:
            process_the_file.COUNT_GROUPS_IN_ONE_PART = 3
            write(sort_by, files, result_filename)
        else:
            process_the_file.COUNT_GROUPS_IN_ONE_PART = 2
            write(sort_by, files, result_filename)

    def get_sheets_title(sheet):
        return sheet.title

    none_sort_wb = openpyxl.load_workbook(f'output/{result_filename}.xlsx')
    none_sort_wb._sheets.sort(key=get_sheets_title)
    none_sort_wb.save(f'output/{result_filename}.xlsx')

########################################################################

def copy_cells_from_to(ws, from_start, from_end, to_start):
    start_line = int(from_start[1])
    end_line = int(from_end[1])
    start_letter = ord(from_start[0])
    end_letter = ord(from_end[0])
    for line in range(start_line, end_line+1):
        for letter in range(start_letter, end_letter+1):
            ws[f'{chr(ord(to_start[0]) + letter - start_letter)}{int(to_start[1:]) + line - start_line}'] = ws[f'{chr(letter)}{line}'].value
            ws[f'{chr(ord(to_start[0]) + letter - start_letter)}{int(to_start[1:]) + line - start_line}']._style = ws[f'{chr(letter)}{line}']._style
    ws.merge_cells(f'B{int(to_start[1:]) + line - start_line - 1}:I{int(to_start[1:]) + line - start_line - 1}')

def write_IiPPO_auditory(sort_by, names, result_filename):
    # preparated_data = get_by(sort_by, process(name))
    preparated_data = {}
    for name in names:
        zagotovka = get_by(sort_by, process_exams(name))
        for k, v in zagotovka.items():
            if preparated_data.get(k) == None:
                preparated_data[k] = []
            preparated_data[k] += v
    preparated_data = dict(sorted(preparated_data.items()))
    try:
        wb = openpyxl.load_workbook(f'output/{result_filename}.xlsx')
    except:
        wb = openpyxl.Workbook()

        # Удаление листа, создаваемого по умолчанию, при создании документа
        for sheet_name in wb.sheetnames:
            sheet = wb.get_sheet_by_name(sheet_name)
            wb.remove_sheet(sheet)

    template = openpyxl.load_workbook('templates/pattern_auditory.xlsx')['Шаблон']

    try:
        ws = wb.get_sheet_by_name('Занятость аудиторий')
    except:
        ws = wb.create_sheet('Занятость аудиторий')
        copy_sheet.copy_sheet(template, ws)

    header = ['A1', 'I2']
    header_line = 1
    line_skip = 2
    last_auditories_count = 0
    fill = {
        'Экзамен': '00BFFF',
        'Консультация': 'B0E0E6',
        'Курсовая работа': '6495ED',
        'Зачёт': '87CEFA',
        'Зачёт/Зачёт': '87CEFA'
    }

    for key, val in preparated_data.items():
        auditories = ['Г-226-1', 'Г-226-2', 'Г-227-1', 'Г-227-2']
        if last_auditories_count != 0:
            header_line += line_skip + last_auditories_count + 2
            copy_cells_from_to(ws, header[0], header[1], f'A{header_line}')
        ws[f'B{header_line}'] = key + ' января'

        for auditory in auditories:
            ws[f'A{auditories.index(auditory) + header_line + 2}'] = auditory

        for info in val:
            if type(info.get('Аудитория')) != float:
                if info.get('Аудитория') not in auditories and info.get('Аудитория').find('\n') == -1:
                    auditories.append(info.get('Аудитория'))
                if info.get('Аудитория').find('\n') == -1:
                    ws[f'A{auditories.index(info.get("Аудитория")) + header_line + 2}'] = info.get('Аудитория')
                    ws[f'{chr(ord("A") + get_parity(info.get("Время"))[0])}{auditories.index(info.get("Аудитория")) + header_line + 2}'] = info.get('Вид занятий')
                    ws[f'{chr(ord("A") + get_parity(info.get("Время"))[0])}{auditories.index(info.get("Аудитория")) + header_line + 2}'].fill = \
                        PatternFill(start_color=fill.get(info.get('Вид занятий')), end_color=fill.get(info.get('Вид занятий')), fill_type="solid")
                    # ws[f'J{line_num}'].fill = PatternFill(start_color="ff0000", end_color="ff0000", fill_type="solid")
                else:
                    print(info)
                    predmets = info.get('Аудитория').split('\n')
                    for i, predmet in enumerate(predmets):
                        if predmet not in auditories:
                            auditories.append(predmet)
                        ws[f'A{auditories.index(predmet) + header_line + 2}'] = predmet
                        if type(info.get('Время')) != float:
                            _time = info.get('Время').split('\n')
                            ws[f'{chr(ord("A") + get_parity(_time[i])[0])}{auditories.index(predmet) + header_line + 2}'] = info.get('Вид занятий')
                            ws[f'{chr(ord("A") + get_parity(_time[i])[0])}{auditories.index(predmet) + header_line + 2}'].fill = \
                                PatternFill(start_color=fill.get(info.get('Вид занятий')), end_color=fill.get(info.get('Вид занятий')), fill_type="solid")
                        else:
                            ws[f'{chr(ord("A") + 7)}{auditories.index(predmet) + header_line + 2}'] = info.get('Вид занятий')
                            ws[f'{chr(ord("A") + 7)}{auditories.index(predmet) + header_line + 2}'].fill = \
                                PatternFill(start_color=fill.get(info.get('Вид занятий')), end_color=fill.get(info.get('Вид занятий')), fill_type="solid")
        last_auditories_count = len(auditories)

    wb.save(f'output/{result_filename}.xlsx')

# write('IIT_3-kurs_22_23_osen_07.10.2022.xlsx')


def write_all_auditories():
    for files in os.listdir('input'):
        if files.find('ITU_mag') != -1:
            process_the_file.COUNT_GROUPS_IN_ONE_PART = 3
            write_auditory(files)
        else:
            process_the_file.COUNT_GROUPS_IN_ONE_PART = 2
            write_auditory(files)

    def get_sheets_title(sheet):
        return sheet.title

    none_sort_wb = openpyxl.load_workbook('output/auditory_result.xlsx')
    none_sort_wb._sheets.sort(key=get_sheets_title)
    none_sort_wb.save('output/auditory_result.xlsx')

########################################################################

def write_lesson(name):
    preparated_data_prepods = process_prepods(name)
    preparated_data = process_the_file.convert_prepods_to(preparated_data_prepods, 'Предмет')
    # pprint.pprint(preparated_data)
    try:
        wb = openpyxl.load_workbook('output/lesson_result.xlsx')
    except:
        wb = openpyxl.Workbook()

        # Удаление листа, создаваемого по умолчанию, при создании документа
        for sheet_name in wb.sheetnames:
            sheet = wb.get_sheet_by_name(sheet_name)
            wb.remove_sheet(sheet)

    template = openpyxl.load_workbook('templates/pattern_prepod.xlsx')['Шаблон']

    for auditory_name, prepods in preparated_data.items():
        # print(auditory_name, prepods)
        if type(auditory_name) != float:
            k = auditory_name.replace('/', ' ')
            try:
                ws = wb.get_sheet_by_name(k)
            except:
                ws = wb.create_sheet(k)
                copy_sheet.copy_sheet(template, ws)
            for prepod_content in prepods:
                for fio, v in prepod_content.items():
                    for weekday, lessons in v.items():
                        for num_lesson, chetnost in lessons.items():
                            for num_chetnost, info in chetnost.items():
                                # print(k, weekday, num_lesson, num_chetnost, info)
                                line_num = 4+weekdays.get(weekday)*16+(int(num_lesson)-1)*2+parity.get(num_chetnost)
                                ws[f'E{line_num}'] = info.get('Предмет')
                                ws[f'F{line_num}'] = info.get('Вид занятий')
                                ws[f'G{line_num}'] = fio
                                ws[f'H{line_num}'] = info.get('Аудитория')
                                ws[f'I{line_num}'] = info.get('Группа')
    wb.save('output/lesson_result.xlsx')

# write('IIT_3-kurs_22_23_osen_07.10.2022.xlsx')


def write_all_lessons():
    for files in os.listdir('input'):
        if files.find('ITU_mag') != -1:
            process_the_file.COUNT_GROUPS_IN_ONE_PART = 3
            write_lesson(files)
        else:
            process_the_file.COUNT_GROUPS_IN_ONE_PART = 2
            write_lesson(files)

    def get_sheets_title(sheet):
        return sheet.title

    none_sort_wb = openpyxl.load_workbook('output/lesson_result.xlsx')
    none_sort_wb._sheets.sort(key=get_sheets_title)
    none_sort_wb.save('output/lesson_result.xlsx')
