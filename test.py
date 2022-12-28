import os
import pprint
import openpyxl
import pandas as pd
import re

# file = pd.read_csv('csv/IIT_mag_2kurs_osen_2022_2023.csv')
# dic = {}
# print(
#     'Шошников И.К.\n\nЧучаева С.М.'.split('\n\n')
# )
#
# pprint.pprint(dic)
# for files in os.listdir('input'):
#     print(files)
#
# print(file.columns[5])

# def get_sheets_title(workbook):
#     return workbook.title
#
# none_sort_wb = openpyxl.load_workbook('output/prepodi_result.xlsx')
# print(none_sort_wb._sheets[0].title)
# none_sort_wb._sheets.sort(key=get_sheets_title)
# none_sort_wb.save('output/prepodi_result.xlsx')

# d = {'a': 21, 'b': 2135}
#
# del d['b']
#
# print(d)
###################################
# from process_the_file import *
#
#
# def find_first_upper_symbol(string: str):
#     for i, char in enumerate(string):
#         if char.isupper():
#             return i
#     return -1
#
# s = 'ASdsdАрхитектура интеграции и развертывания'
# print(s.split('\n'))
# print(s[find_first_upper_symbol(s):])
#
# aud = 'ауд. G-234 (B-78)'
# aud1 = 'комп. G-234 (B-78)'
# aud0 = 'G-234'
#
# def get_clear_auditory(auditory: str):
#     if auditory.find('(') != -1:
#         auditory = auditory[:auditory.find('(')-1]
#     if auditory.find('ауд. ') != -1:
#         auditory = auditory[auditory.find('ауд. ')+5:]
#     if auditory.find('комп. ') != -1:
#         auditory = auditory[auditory.find('комп. ')+6:]
#     return auditory
#
# print(get_clear_auditory(aud))
# print(get_clear_auditory(aud0))
# print(get_clear_auditory(aud1))
####################################
# def edit_d(d):
#     d['a'] = 124
#
# edit_d(d)
# print(d)

# fios = 'Глухов А.В.' \
# 'Прилипко В.А.' \
# 'Прилипко В.А.'

# file = pd.read_csv('csv/IIT_mag_2kurs_osen_2022_2023.csv')
# fios = file.iloc[59, 72]
#
# print(fios.split('\n'))

# for files in os.listdir('input'):
#     print(files)
# pprint.pprint(process_the_file.get_by("ФИО", process_the_file.process('зач_ИИТ_3 курс_22-23_осень.csv')))

# print(process_the_file.get_clear_fio("Пяткин В.В."))
# print("Пяткин В.В."[0:1])

# def split_fios(fio):
#     while get_n_upper_symbol(fio) > 0:
#         print(fio[:find_n_upper_symbol(fio, 2) + 2])
#         fio = fio[find_n_upper_symbol(fio, 2) + 2:]
#
# split_fios("Мильчакова Н.Е.")

# print("Мильчакова Н.Е.\nМочаловаЛ.В.")

# ban_symbols = [',', '\n\n', '\n', ', ']
#
# if not any(ban_symbol in "Мильчакова Н.Е.\n\nМочаловаЛ.В." for ban_symbol in ban_symbols):
#     print('Success')

# s = 'ИКБО-30-20'
#
# pattern = re.compile("[А-Я]{4}-[0-9]{2}-[0-9]{2}")
# res = pattern.search(s)
#
# print(type(pattern))
# print(res)

# file = pd.read_csv('csv/IIT_1-kurs_2022_2023_zima (1).csv')
# fios = file.iloc[0, :10]
#
# print(fios.iloc[10, :])



import process_the_file
import DateTime

# file = process_the_file.process_exams('IIT_1-kurs_2022_2023_zima (1).xlsx')
# data = process_the_file.get_by('ФИО', file)
#
# times = []
# # times.append('20-00')
#
# def sort_time(time):
#     return DateTime.DateTime(time)
#
# for el in file:
#     if el.get('Время') not in times:
#         times.append(el.get('Время'))
#
# print(sorted(times, key=sort_time))
# # pprint.pprint(file)
# pprint.pprint(data)



# parity_num_time = {
#     1: '9-00',
#     2: '10-40',
#     3: '12-40',
#     4: '14-20',
#     5: '16-20',
#     6: '18-00'
# }
#
# def get_parity(time):
#     time = DateTime.DateTime(time)
#     for k, v in parity_num_time.items():
#         if time <= DateTime.DateTime(v):
#             if k > 1:
#                 if DateTime.DateTime(v) - time < time - DateTime.DateTime(parity_num_time.get(k-1)):
#                     return (k, v)
#                 else:
#                     return (k-1, parity_num_time.get(k-1))
#             else:
#                 if time - DateTime.DateTime(v) < DateTime.DateTime(parity_num_time.get(k+1)) - time:
#                     return (k, v)
#                 else:
#                     return (k+1, parity_num_time.get(k+1))
#
# print((get_parity('10-00')[0]-1))

# print(chr(65+4))
#
# for i in range(ord('A'), ord('I')+1):
#     print(chr(i))

# main_auditories = ['Г-226-1', 'Г-226-2', 'Г-227-1', 'Г-227-2']
# print(main_auditories.index('Г-227-1'))

# def copy_cells_from_to(from_start, from_end, to_start):
#     start_line = int(from_start[1])
#     end_line = int(from_end[1])
#     start_letter = ord(from_start[0])
#     end_letter = ord(from_end[0])
#     for line in range(start_line, end_line+1):
#         for letter in range(start_letter, end_letter+1):
#             # ws[f'{chr(ord(to_start[0]) + letter - start_letter)}{int(to_start[1]) + line - start_line}'] = ws[f'{chr(letter)}{line}']
#             # print(f'{chr(letter)}{line}')
#             print( f'{chr(ord(to_start[0]) + letter - start_letter)}{int(to_start[1]) + line - start_line}')
#             # print( f'{chr(ord(to_start[0]) + letter - start_letter)}{int(to_start[1]) + line - start_line}' + " ::: " + f'{chr(letter)}{line}')
#
# copy_cells_from_to('A1', 'C3', 'A5')


import copy_sheet

def copy_cells_from_to(ws, from_start, from_end, to_start):
    start_line = int(from_start[1])
    end_line = int(from_end[1])
    start_letter = ord(from_start[0])
    end_letter = ord(from_end[0])
    for line in range(start_line, end_line+1):
        for letter in range(start_letter, end_letter+1):
            print(f'{chr(ord(to_start[0]) + letter - start_letter)}{int(to_start[1:]) + line - start_line}' + ' => ' + f'{chr(letter)}{line}')
            print(f'{line} {start_line}')
            ws[f'{chr(ord(to_start[0]) + letter - start_letter)}{int(to_start[1:]) + line - start_line}'] = ws[f'{chr(letter)}{line}'].value

def write_IiPPO_auditory(result_filename):
    # preparated_data = get_by(sort_by, process(name))
    try:
        wb = openpyxl.load_workbook(f'output/{result_filename}.xlsx')
    except:
        wb = openpyxl.Workbook()

        # Удаление листа, создаваемого по умолчанию, при создании документа
        for sheet_name in wb.sheetnames:
            sheet = wb.get_sheet_by_name(sheet_name)
            wb.remove_sheet(sheet)

    template = openpyxl.load_workbook('templates/pattern_test.xlsx')['Шаблон']

    try:
        ws = wb.get_sheet_by_name('Занятость аудиторий')
    except:
        ws = wb.create_sheet('Занятость аудиторий')
        copy_sheet.copy_sheet(template, ws)

    header = ['A1', 'I2']
    header_line = 1
    line_skip = 5
    last_auditories_count = 0
    for i in range(2):
        # auditories = ['Г-226-1', 'Г-226-2', 'Г-227-1', 'Г-227-2']
        header_line += line_skip + last_auditories_count + 1
        copy_cells_from_to(ws, header[0], header[1], f'A{header_line}')
        print(header_line)
        ws[f'B{header_line}'] = str(i) + ' января'

    wb.save(f'output/{result_filename}.xlsx')

write_IiPPO_auditory('test')